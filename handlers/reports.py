import logging
from datetime import date, datetime

from telegram import Update
from telegram.ext import ContextTypes

from database import get_collection
from keyboards import menu_reportes
from utils import calcular_dias_vencido

logger = logging.getLogger(__name__)


async def menu_reports(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("📊 Menu reportes", reply_markup=menu_reportes)


async def deudores(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    members = get_collection("members")
    payments = get_collection("payments")

    hoy = date.today()
    texto = "⚠️ MIEMBROS CON PAGOS VENCIDOS:\n\n"

    all_members = list(members.find({"active": True}))

    if not all_members:
        await update.message.reply_text("No hay miembros registrados")
        return

    deudores_count = 0

    for member in all_members:
        last_payment = payments.find_one(
            {"member_id": str(member["_id"])},
            sort=[("payment_date", -1)]
        )

        if not last_payment:
            continue

        due_date = datetime.strptime(last_payment["due_date"], "%Y-%m-%d").date()
        dias_vencido = calcular_dias_vencido(due_date)

        if dias_vencido == 0:
            if hoy == due_date:
                texto += f"• {member['name']}\n"
                texto += f"  ⏰ Vence hoy: {last_payment['due_date']}\n\n"
                deudores_count += 1
            continue
        elif 1 <= dias_vencido <= 4:
            texto += f"• {member['name']}\n"
            texto += f"  ⚠️ En gracia ({dias_vencido} dias)\n"
            texto += f"  Vencio: {last_payment['due_date']}\n\n"
            deudores_count += 1
        else:
            texto += f"• {member['name']}\n"
            texto += f"  💀 Vencio: {last_payment['due_date']}\n"
            texto += f"  📅 Dias vencido: {dias_vencido}\n\n"
            deudores_count += 1

    if deudores_count == 0:
        texto = "✅ Todos los miembros estan al dia"

    await update.message.reply_text(texto)


async def excel_reporte(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    import os

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    from config import EXCEL_FILE

    members = get_collection("members")
    payments = get_collection("payments")

    os.makedirs("reports", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Miembros"

    ws.append(["Nombre", "Fecha Registro", "Ultimo Pago", "Vence", "Plan", "Dias Vencido", "Estado"])

    verde = PatternFill(start_color="90EE90", fill_type="solid")
    rojo = PatternFill(start_color="FF7F7F", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF99", fill_type="solid")

    hoy = date.today()

    all_members = list(members.find({"active": True}))

    for member in all_members:
        last_payment = payments.find_one(
            {"member_id": str(member["_id"])},
            sort=[("payment_date", -1)]
        )

        if last_payment:
            vencimiento = datetime.strptime(last_payment["due_date"], "%Y-%m-%d").date()
            dias_vencido = (hoy - vencimiento).days

            if dias_vencido < 0:
                estado = "Al dia"
                dias_display = 0
                fill = verde
            elif dias_vencido == 0:
                estado = "Vence hoy"
                dias_display = 0
                fill = amarillo
            elif dias_vencido <= 4:
                estado = "En gracia"
                dias_display = dias_vencido
                fill = amarillo
            else:
                estado = "Vencido"
                dias_display = dias_vencido
                fill = rojo

            ws.append([
                member["name"],
                member["created_at"].strftime("%Y-%m-%d"),
                last_payment["payment_date"],
                last_payment["due_date"],
                last_payment["plan"],
                dias_display,
                estado,
            ])

            fila = ws.max_row
            ws[f"G{fila}"].fill = fill

    wb.save(EXCEL_FILE)

    with open(EXCEL_FILE, "rb") as f:
        await update.message.reply_document(
            f,
            filename="reporte_miembros.xlsx"
        )
