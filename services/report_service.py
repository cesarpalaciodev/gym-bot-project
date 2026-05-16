from __future__ import annotations

import logging
import os
from collections.abc import Callable
from datetime import date, datetime
from typing import Any

from motor.motor_asyncio import AsyncIOMotorCollection
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from utils import calcular_dias_vencido

logger = logging.getLogger(__name__)


class ReportService:
    def __init__(
        self,
        members_col: AsyncIOMotorCollection[Any],
        payments_col: AsyncIOMotorCollection[Any],
        *,
        today_fn: Callable[[], date] | None = None,
    ) -> None:
        self.members = members_col
        self.payments = payments_col
        self._today = today_fn or date.today

    async def get_members_with_payments(self) -> list[tuple[dict[str, Any], dict[str, Any] | None]]:
        all_members = await self.members.find({"active": True}).to_list(None)
        result: list[tuple[dict[str, Any], dict[str, Any] | None]] = []
        for member in all_members:
            last_payment = await self.payments.find_one({"member_id": str(member["_id"])}, sort=[("payment_date", -1)])
            result.append((member, last_payment))
        return result

    async def get_income_data(self) -> dict[str, Any]:
        hoy = self._today()
        inicio_mes = hoy.replace(day=1)
        from dateutil.relativedelta import relativedelta

        mes_actual = await self.payments.find(
            {"payment_date": {"$gte": inicio_mes.strftime("%Y-%m-%d"), "$lte": hoy.strftime("%Y-%m-%d")}}
        ).to_list(None)
        inicio_mes_pasado = inicio_mes - relativedelta(months=1)
        fin_mes_pasado = inicio_mes - relativedelta(days=1)
        mes_pasado = await self.payments.find(
            {
                "payment_date": {
                    "$gte": inicio_mes_pasado.strftime("%Y-%m-%d"),
                    "$lte": fin_mes_pasado.strftime("%Y-%m-%d"),
                }
            }
        ).to_list(None)

        monto_actual = sum(p["amount"] for p in mes_actual)
        monto_pasado = sum(p["amount"] for p in mes_pasado)
        return {
            "current_month_payments": mes_actual,
            "previous_month_payments": mes_pasado,
            "current_amount": monto_actual,
            "previous_amount": monto_pasado,
        }

    async def get_overdue_text(self) -> str:
        hoy = self._today()
        texto = "\u26a0\ufe0f MIEMBROS CON PAGOS VENCIDOS:\n\n"
        all_members = await self.members.find({"active": True}).to_list(None)
        if not all_members:
            return "No hay miembros registrados"

        deudores_count = 0
        for member in all_members:
            last_payment = await self.payments.find_one({"member_id": str(member["_id"])}, sort=[("payment_date", -1)])
            if not last_payment:
                continue
            due_date = datetime.strptime(last_payment["due_date"], "%Y-%m-%d").date()
            dias_vencido = calcular_dias_vencido(due_date)

            if dias_vencido == 0:
                if hoy == due_date:
                    texto += f"\u2022 {member['name']}\n"
                    texto += f"  \u23f0 Vence hoy: {last_payment['due_date']}\n\n"
                    deudores_count += 1
                continue
            elif 1 <= dias_vencido <= 4:
                texto += f"\u2022 {member['name']}\n"
                texto += f"  \u26a0\ufe0f En gracia ({dias_vencido} dias)\n"
                texto += f"  Vencio: {last_payment['due_date']}\n\n"
                deudores_count += 1
            else:
                texto += f"\u2022 {member['name']}\n"
                texto += f"  \U0001f480 Vencio: {last_payment['due_date']}\n"
                texto += f"  \U0001f4c5 Dias vencido: {dias_vencido}\n\n"
                deudores_count += 1

        if deudores_count == 0:
            texto = "\u2705 Todos los miembros estan al dia"
        return texto

    async def generate_excel(self, filepath: str) -> str:
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Miembros"
        ws.append(["Nombre", "Fecha Registro", "Ultimo Pago", "Vence", "Plan", "Dias Vencido", "Estado"])

        verde = PatternFill(start_color="90EE90", fill_type="solid")
        rojo = PatternFill(start_color="FF7F7F", fill_type="solid")
        amarillo = PatternFill(start_color="FFFF99", fill_type="solid")
        hoy = self._today()
        all_members = await self.members.find({"active": True}).to_list(None)

        for member in all_members:
            last_payment = await self.payments.find_one({"member_id": str(member["_id"])}, sort=[("payment_date", -1)])
            if not last_payment:
                continue
            vencimiento = datetime.strptime(last_payment["due_date"], "%Y-%m-%d").date()
            dias_vencido = (hoy - vencimiento).days
            if dias_vencido < 0:
                estado, dias_display, fill = "Al dia", 0, verde
            elif dias_vencido == 0:
                estado, dias_display, fill = "Vence hoy", 0, amarillo
            elif dias_vencido <= 4:
                estado, dias_display, fill = "En gracia", dias_vencido, amarillo
            else:
                estado, dias_display, fill = "Vencido", dias_vencido, rojo
            ws.append(
                [
                    member["name"],
                    member.get("created_at", datetime.utcnow()).strftime("%Y-%m-%d"),
                    last_payment["payment_date"],
                    last_payment["due_date"],
                    last_payment["plan"],
                    dias_display,
                    estado,
                ]
            )
            ws[f"G{ws.max_row}"].fill = fill

        wb.save(filepath)
        return filepath
