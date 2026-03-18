import json
import os
import logging
import shutil
from datetime import datetime, time
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
load_dotenv()
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from supabase import create_client

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
if not SUPABASE_URL or not SUPABASE_KEY:
    
    raise ValueError("Faltan variables de entorno de Supabase")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# ==================================================
# CONFIGURACION
# ==================================================


TOKEN = os.getenv("TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

DATA_FILE = "/var/data/miembros.json"
EXCEL_FILE = "reports/reporte_gimnasio.xlsx"
BACKUP_FOLDER = "backup"

# ==================================================
# LOGGING
# ==================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# ==================================================
# ESTADO USUARIO
# ==================================================

user_state = {}

# ==================================================
# UTILIDADES
# ==================================================

def es_admin(update):
    return update.effective_user.id == ADMIN_ID


def cargar_datos():
    if not os.path.exists(DATA_FILE):
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f)
        return {}

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def guardar_datos(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)
def guardar_miembro(nombre, fecha):
    data = {
        "nombre": nombre,
        "fecha": fecha
    }

    response = supabase.table("miembros").insert(data).execute()

    print(response)

def fecha_valida(fecha):
    try:
        datetime.strptime(fecha, "%Y-%m-%d")
        return True
    except:
        return False

# ==================================================
# BACKUP
# ==================================================

async def backup(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if not es_admin(update):
        await update.message.reply_text("No autorizado")
        return

    if not os.path.exists(BACKUP_FOLDER):
        os.makedirs(BACKUP_FOLDER)

    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
    archivo_backup = f"{BACKUP_FOLDER}/miembros_backup_{fecha}.json"

    shutil.copy(DATA_FILE, archivo_backup)

    await update.message.reply_text("✅ Backup creado correctamente")

# ==================================================
# BACKUP AUTOMATICO
# ==================================================

async def backup_automatico(context: ContextTypes.DEFAULT_TYPE):

    if not os.path.exists(BACKUP_FOLDER):
        os.makedirs(BACKUP_FOLDER)

    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
    archivo_backup = f"{BACKUP_FOLDER}/miembros_backup_{fecha}.json"

    shutil.copy(DATA_FILE, archivo_backup)

# ==================================================
# EXCEL
# ==================================================

def generar_excel():

    data = cargar_datos()

    wb = Workbook()
    ws = wb.active

    ws.append(["Nombre", "Vencimiento", "Estado"])

    verde = PatternFill(start_color="90EE90", fill_type="solid")
    rojo = PatternFill(start_color="FF7F7F", fill_type="solid")

    hoy = datetime.now().date()

    for nombre, fecha in data.items():

        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        vencimiento = fecha_dt + relativedelta(months=1)

        estado = "Al día"

        if hoy >= vencimiento:
            estado = "Vencido"

        ws.append([nombre, vencimiento.strftime("%Y-%m-%d"), estado])

        fila = ws.max_row

        if estado == "Al día":
            ws[f"C{fila}"].fill = verde
        else:
            ws[f"C{fila}"].fill = rojo

    wb.save(EXCEL_FILE)

# ==================================================
# MENUS
# ==================================================

menu_principal = ReplyKeyboardMarkup(
[
["👥 Miembros","💰 Pagos"],
["📊 Reportes"]
],
resize_keyboard=True
)

menu_miembros = ReplyKeyboardMarkup(
[
["➕ Agregar miembro","👥 Agregar varios"],
["🔍 Buscar miembro","📋 Lista miembros"],
["🗑 Eliminar miembro","🗑 Eliminar varios"],
["⬅️ Volver"]
],
resize_keyboard=True
)

menu_pagos = ReplyKeyboardMarkup(
[
["💰 Registrar pago"],
["⬅️ Volver"]
],
resize_keyboard=True
)

menu_reportes = ReplyKeyboardMarkup(
[
["⚠️ Deudores","📊 Excel"],
["⬅️ Volver"]
],
resize_keyboard=True
)

# ==================================================
# START
# ==================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if not es_admin(update):
        await update.message.reply_text("Acceso no autorizado")
        return

    await update.message.reply_text(
        "🏋️ Sistema del gimnasio",
        reply_markup=menu_principal
    )

# ==================================================
# FUNCIONES GYM
# ==================================================

async def lista(update):

    data = cargar_datos()

    if not data:
        await update.message.reply_text("No hay miembros")
        return

    texto = "Miembros:\n\n"

    for n, f in data.items():
        texto += f"{n} - {f}\n"

    await update.message.reply_text(texto)


async def deudores(update):

    data = cargar_datos()
    hoy = datetime.now().date()

    texto = "Deudores:\n\n"

    for n, f in data.items():

        fecha = datetime.strptime(f, "%Y-%m-%d").date()
        vencimiento = fecha + relativedelta(months=1)

        if hoy >= vencimiento:
            texto += f"{n} - debía pagar el {vencimiento}\n"

    if texto == "Deudores:\n\n":
        texto = "Todos al día"

    await update.message.reply_text(texto)

# ==================================================
# BOTONES
# ==================================================

async def botones(update: Update, context: ContextTypes.DEFAULT_TYPE):

    texto = update.message.text
    user_id = update.effective_user.id

    if not es_admin(update):
        return

    if texto == "👥 Miembros":
        await update.message.reply_text("Menú miembros",reply_markup=menu_miembros)

    elif texto == "💰 Pagos":
        await update.message.reply_text("Menú pagos",reply_markup=menu_pagos)

    elif texto == "📊 Reportes":
        await update.message.reply_text("Menú reportes",reply_markup=menu_reportes)

    elif texto == "⬅️ Volver":
        await update.message.reply_text("Menú principal",reply_markup=menu_principal)

    elif texto == "📋 Lista miembros":
        await lista(update)

    elif texto == "⚠️ Deudores":
        await deudores(update)

    elif texto == "📊 Excel":
        generar_excel()
        await update.message.reply_document(open(EXCEL_FILE,"rb"))

    elif texto == "➕ Agregar miembro":

        user_state[user_id] = "agregar"

        await update.message.reply_text(
            "Escribe:\nNombre YYYY-MM-DD\nEjemplo:\nCarlos 2026-03-20"
        )

    elif texto == "👥 Agregar varios":

        user_state[user_id] = "varios"

        await update.message.reply_text(
            "Escribe uno por línea:\nCarlos 2026-03-20"
        )

    elif texto == "🔍 Buscar miembro":

        user_state[user_id] = "buscar"

        await update.message.reply_text("Escribe el nombre")

    elif texto == "🗑 Eliminar miembro":

        user_state[user_id] = "eliminar"

        await update.message.reply_text("Nombre a eliminar")

    elif texto == "🗑 Eliminar varios":

        user_state[user_id] = "eliminar_varios"

        await update.message.reply_text("Escribe los nombres uno por línea")

    elif texto == "💰 Registrar pago":

        user_state[user_id] = "pago"

        await update.message.reply_text("Nombre del miembro")

    elif user_id in user_state:

        estado = user_state[user_id]
        data = cargar_datos()

        if estado == "agregar":

            nombre,fecha = texto.split()

            if not fecha_valida(fecha):
                await update.message.reply_text("Fecha inválida")
                return

            data[nombre] = fecha
            guardar_datos(data)
            guardar_miembro(nombre, fecha)

            await update.message.reply_text("Miembro agregado")
            del user_state[user_id]

        elif estado == "varios":

            for l in texto.split("\n"):
                try:
                    nombre,fecha = l.split()
                    data[nombre] = fecha
                except:
                    pass

            guardar_datos(data)

            await update.message.reply_text("Miembros agregados")
            del user_state[user_id]

        elif estado == "buscar":

            if texto in data:

                fecha_pago = datetime.strptime(data[texto], "%Y-%m-%d").date()
                vencimiento = fecha_pago + relativedelta(months=1)

                await update.message.reply_text(
                    f"👤 {texto}\n"
                    f"💰 Último pago: {fecha_pago}\n"
                    f"📅 Vence: {vencimiento}"
                )

            else:
                await update.message.reply_text("No encontrado")

            del user_state[user_id]

        elif estado == "eliminar":

            if texto in data:
                del data[texto]
                guardar_datos(data)
                await update.message.reply_text("Eliminado")
            else:
                await update.message.reply_text("No existe")

            del user_state[user_id]

        elif estado == "eliminar_varios":

            eliminados = []
            no_encontrados = []

            for nombre in texto.split("\n"):

                nombre = nombre.strip()

                if nombre in data:
                    del data[nombre]
                    eliminados.append(nombre)
                else:
                    no_encontrados.append(nombre)

            guardar_datos(data)

            mensaje = "Eliminados:\n"
            for e in eliminados:
                mensaje += f"- {e}\n"

            if no_encontrados:
                mensaje += "\nNo encontrados:\n"
                for n in no_encontrados:
                    mensaje += f"- {n}\n"

            await update.message.reply_text(mensaje)

            del user_state[user_id]

        elif estado == "pago":

            if texto not in data:
                await update.message.reply_text("No existe")
                return

            data[texto] = datetime.now().date().strftime("%Y-%m-%d")

            guardar_datos(data)

            await update.message.reply_text("Pago registrado")

            del user_state[user_id]

# ==================================================
# MAIN
# ==================================================

def main():

    app = Application.builder().token(TOKEN).build()

    job_queue = app.job_queue

    job_queue.run_daily(
         backup_automatico,
         time=time(hour=22, minute=0)
    )

    app.add_handler(CommandHandler("start",start))
    app.add_handler(CommandHandler("backup",backup))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND,botones))

    print("Bot iniciado")

    app.run_polling()

if __name__ == "__main__":
    main()